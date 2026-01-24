# Reconciliation Automation Script – v37.1 AI Integrated with CLEARINGOPER Enhancement
# Author: Qwen + Ozan
# Date: 2026-01-24
# Platform: [Office PC / Jupyter] + Ollama (gemma3:4b)
# Purpose: Process all 5 source files + CLEARINGOPER, perform reconciliation, validate results with AI, and integrate MC ANR.

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import re
import glob # Import glob for file searching
from datetime import datetime # Import datetime for timestamp
import requests # Import requests to interact with Ollama API

# --- CONFIGURATION - Keywords for File Discovery ===
MC_SUMMARY_KEYWORDS = ["IncomingSummary", "MasterCard"]
VISA_SUMMARY_KEYWORDS = ["IncomingSummary", "VISA"]
MC_DPR_EXCEL_KEYWORDS = ["MastercardDPRReport"]
VISA_DPR_TEXT_KEYWORDS = ["NSPK-VISA-VSS"]
MIR_RECI_TEXT_KEYWORDS = ["NSPK-MIR-ANR"]
# NEW: Keyword for CLEARINGOPER file
CLEARINGOPER_KEYWORDS = ["CLEARINGOPER"]

# Modified Template File Name
TEMPLATE_FILE_NAME = "Reconciliation Process 01 - Bank vs PS - Template.xlsx"

# === OLLAMA Configuration ===
OLLAMA_URL = "http://localhost:11434/api/generate" # Default Ollama URL
OLLAMA_MODEL = "gemma3:4b" # Model to use

# === Helper Function to Find File by Keywords ===
def find_file_by_keywords(keywords):
    search_pattern = "*"
    for keyword in keywords:
        search_pattern += f"*{keyword}*"
    search_pattern += "*"

    matching_files = glob.glob(search_pattern)
    matching_files = [f for f in matching_files if os.path.isfile(f)]

    if len(matching_files) == 0:
        raise FileNotFoundError(f"No file found containing all keywords: {keywords}")
    elif len(matching_files) > 1:
        raise FileNotFoundError(f"Multiple files found containing keywords {keywords}: {matching_files}. Please ensure uniqueness.")
    else:
        return matching_files[0]

# Generate output filename with timestamp and prefix
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# Modified Output File Names - Added "Result - " prefix
output_file = f"Result - Reconciliation_Auto_{timestamp}.xlsx"
ai_validation_output_file = f"Result - AI_Reconciliation_Report_{timestamp}.txt" # File to save AI validation analysis
# Note: Chart code file will be handled separately to avoid blocking
chart_code_output_file = f"Result - neon_chart_code_{timestamp}.py" # File to save AI-generated chart code if successful

# === Locate Files Dynamically ===
try:
    mc_summary_file = find_file_by_keywords(MC_SUMMARY_KEYWORDS)
    visa_summary_file = find_file_by_keywords(VISA_SUMMARY_KEYWORDS)
    dpr_mc_file = find_file_by_keywords(MC_DPR_EXCEL_KEYWORDS)
    dpr_visa_file = find_file_by_keywords(VISA_DPR_TEXT_KEYWORDS)
    dpr_mir_file = find_file_by_keywords(MIR_RECI_TEXT_KEYWORDS)
    clearingoper_file = find_file_by_keywords(CLEARINGOPER_KEYWORDS)  # NEW: Find CLEARINGOPER file
    template_file = TEMPLATE_FILE_NAME
    
    print(f"--- File Discovery Results ---")
    print(f"MC Summary: {mc_summary_file}")
    print(f"VISA Summary: {visa_summary_file}")
    print(f"MC DPR Excel: {dpr_mc_file}")
    print(f"VISA DPR Text: {dpr_visa_file}")
    print(f"MIR RECI Text: {dpr_mir_file}")
    print(f"CLEARINGOPER File: {clearingoper_file}")  # NEW: Print CLEARINGOPER file
    print(f"Template: {template_file}")
    print(f"Output: {output_file}")
    print(f"---------------------------")
    
except FileNotFoundError as e:
    print(f"Error finding files: {e}")
    exit(1)

# === STEP 1: Load Incoming Summary Files ===
df_mc = pd.read_excel(mc_summary_file, sheet_name="Sheet1")
df_visa = pd.read_excel(visa_summary_file, sheet_name="Sheet1")

df_mc.columns = df_mc.columns.str.strip()
df_visa.columns = df_visa.columns.str.strip()

def compute_subtotals(df, system_name):
    data = df[df["Payment System"] == system_name]
    debit_amt = data[data["Credit/Debit Amount Flag"] == "D"]["Total Settlement Amount"].sum()
    credit_amt = data[data["Credit/Debit Amount Flag"] == "C"]["Total Settlement Amount"].sum()
    debit_cnt = data[data["Credit/Debit Amount Flag"] == "D"]["Count"].sum()
    credit_cnt = data[data["Credit/Debit Amount Flag"] == "C"]["Count"].sum()
    return credit_amt, credit_cnt, debit_amt, debit_cnt

c_amt_mc, c_cnt_mc, d_amt_mc, d_cnt_mc = compute_subtotals(df_mc, "NSPK-MC")
net_amt_mc = d_amt_mc - c_amt_mc
net_cnt_mc = d_cnt_mc + c_cnt_mc

c_amt_visa, c_cnt_visa, d_amt_visa, d_cnt_visa = compute_subtotals(df_visa, "NSPK-VISA")
net_amt_visa = d_amt_visa - c_amt_visa
net_cnt_visa = d_cnt_visa + c_cnt_visa

c_amt_mir, c_cnt_mir, d_amt_mir, d_cnt_mir = compute_subtotals(df_mc, "MIR")
net_amt_mir = d_amt_mir - c_amt_mir
net_cnt_mir = d_cnt_mir + c_cnt_mir

# --- Derive VISA VSS values from Incoming Summary and apply sign flip ---
visa_net_amount_for_e20 = -net_amt_visa
visa_count = net_cnt_visa

# === STEP 2: Parse Cleaned MIR RECI File ===
mir_net_amount = 0.0
mir_count = 0

with open(dpr_mir_file, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.rstrip()
        tokens = line.split()
        if len(tokens) < 13:
            continue
        try:
            count_val = int(float(tokens[7]))
            amount_val = float(tokens[11].replace(',', ''))
            dc_flag = tokens[12]
            if dc_flag == "D":
                mir_net_amount -= amount_val
            elif dc_flag == "C":
                mir_net_amount += amount_val
            else:
                continue
            mir_count += count_val
        except (ValueError, IndexError):
            continue

mir_net_amount = round(mir_net_amount, 2)

# === STEP 3: Load DPR Excel for MC NPSK Only ===
df_dpr_mc = pd.read_excel(dpr_mc_file, sheet_name="Sheet1")
df_dpr_mc.columns = df_dpr_mc.columns.str.strip()

dpr_mc_filtered = df_dpr_mc[
    (df_dpr_mc["Clearing System"] == "MC NPSK") &
    (df_dpr_mc["Direction"] == "Incoming")
]
dpr_mc_count = dpr_mc_filtered["Count"].sum()
dpr_mc_amount = dpr_mc_filtered["Net Recon Amount"].sum()

# === NEW: STEP 4: Process CLEARINGOPER File ===
# Load the CLEARINGOPER file
df_clearingoper = pd.read_excel(clearingoper_file)
df_clearingoper.columns = df_clearingoper.columns.str.strip()

# Find the 'ProcessedRecords' column (could be in any position)
processed_records_col = None
for col in df_clearingoper.columns:
    if 'ProcessedRecords' in str(col):
        processed_records_col = col
        break

# Find the 'Flowcode' column (typically column B, but we'll search for it anyway)
flowcode_col = None
for i, col in enumerate(df_clearingoper.columns):
    if i == 1 or 'Flowcode' in str(col) or 'flowcode' in str(col) or 'FLOWCODE' in str(col):
        flowcode_col = col
        break

# Calculate the required counts
miri_count = 0
nvi_count = 0
nmi_count = 0

if processed_records_col and flowcode_col:
    # Count records where Flowcode is MIRI
    miri_df = df_clearingoper[df_clearingoper[flowcode_col] == 'MIRI']
    if not miri_df.empty and processed_records_col in miri_df.columns:
        miri_count = miri_df[processed_records_col].sum()
    
    # Count records where Flowcode is NVI
    nvi_df = df_clearingoper[df_clearingoper[flowcode_col] == 'NVI']
    if not nvi_df.empty and processed_records_col in nvi_df.columns:
        nvi_count = nvi_df[processed_records_col].sum()
    
    # Count records where Flowcode is NMI
    nmi_df = df_clearingoper[df_clearingoper[flowcode_col] == 'NMI']
    if not nmi_df.empty and processed_records_col in nmi_df.columns:
        nmi_count = nmi_df[processed_records_col].sum()

print(f"--- CLEARINGOPER Processing Results ---")
print(f"MIRI count (for F26): {miri_count}")
print(f"NVI count (for F25): {nvi_count}")
print(f"NMI count (for F24): {nmi_count}")
print(f"---------------------------")

# === STEP 5: Load Template and Populate Initial Cells (Up to E21) ===
df_target = pd.read_excel(template_file, header=None)

while len(df_target) < 22:
    df_target = pd.concat([df_target, pd.DataFrame([[""] * df_target.shape[1]], columns=df_target.columns)], ignore_index=True)

df_target.iloc[7, 4] = c_amt_mc      # E8
df_target.iloc[7, 5] = c_cnt_mc
df_target.iloc[8, 4] = d_amt_mc      # E9
df_target.iloc[8, 5] = d_cnt_mc
df_target.iloc[9, 4] = -net_amt_mc   # E10
df_target.iloc[9, 5] = net_cnt_mc

df_target.iloc[10, 4] = c_amt_visa   # E11
df_target.iloc[10, 5] = c_cnt_visa
df_target.iloc[11, 4] = d_amt_visa   # E12
df_target.iloc[11, 5] = d_cnt_visa
df_target.iloc[12, 4] = -net_amt_visa  # E13
df_target.iloc[12, 5] = net_cnt_visa

df_target.iloc[13, 4] = c_amt_mir    # E14
df_target.iloc[13, 5] = c_cnt_mir
df_target.iloc[14, 4] = d_amt_mir    # E15
df_target.iloc[14, 5] = d_cnt_mir
df_target.iloc[15, 4] = -net_amt_mir # E16
df_target.iloc[15, 5] = net_cnt_mir

df_target.iloc[18, 4] = dpr_mc_amount  # E19
df_target.iloc[18, 5] = dpr_mc_count   # F19

df_target.iloc[19, 4] = visa_net_amount_for_e20  # E20
df_target.iloc[19, 5] = visa_count       # F20

df_target.iloc[20, 4] = mir_net_amount  # E21
df_target.iloc[20, 5] = mir_count       # F21

# NEW: Populate F24, F25, F26 with CLEARINGOPER counts
df_target.iloc[23, 5] = nmi_count   # F24 (row 24, col 6)
df_target.iloc[24, 5] = nvi_count   # F25 (row 25, col 6)
df_target.iloc[25, 5] = miri_count  # F26 (row 26, col 6)

# Save intermediate
temp_file = "temp_values_only.xlsx"
df_target.to_excel(temp_file, index=False, header=False)

# === STEP 6: Apply Formatting and Formulas (Up to G26) ===
wb = load_workbook(temp_file)
ws = wb.active

consolas_font = Font(name="Consolas", size=10)

for row in range(1, 27):  # Extended to row 26
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = consolas_font

for row in [8,9,10,11,12,13,14,15,16,19,20,21]:
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=6).number_format = '#,##0'

# Original formulas
ws['G19'] = '=E10-E19'  # MC
ws['G20'] = '=E13-E20'  # VISA
ws['G21'] = '=E16-E21'  # MIR

# NEW: Add formulas for G24, G25, G26 based on requirements
ws['G24'] = '=F22-F24'  # F22-F24 (NMI)
ws['G25'] = '=F20-F25'  # F20-F25 (NVI)
ws['G26'] = '=F21-F26'  # F21-F26 (MIRI)

for row in range(17, 27):  # Extended to row 26
    cell = ws.cell(row=row, column=7)
    cell.number_format = '#,##0.00'
    cell.font = consolas_font

# === STEP 7: Prepare Data for AI Analysis ===
results_summary = {
    "date": timestamp[:8], # Extract YYYYMMDD from timestamp
    "mc_dpr_amount": dpr_mc_amount,
    "mc_dpr_count": dpr_mc_count,
    "visa_vss_amount": visa_net_amount_for_e20,
    "visa_vss_count": visa_count,
    "mir_anr_amount": mir_net_amount,
    "mir_anr_count": mir_count,
    "mc_difference": -net_amt_mc - dpr_mc_amount, # G19 = E10 - E19
    "visa_difference": -net_amt_visa - visa_net_amount_for_e20, # G20 = E13 - E20
    "mir_difference": -net_amt_mir - mir_net_amount, # G21 = E16 - E21
    # NEW: Add CLEARINGOPER values to the summary
    "nmi_count": nmi_count,  # F24
    "nvi_count": nvi_count,  # F25
    "miri_count": miri_count, # F26
    "nmi_formula_result": ws['G24'].value,  # G24
    "nvi_formula_result": ws['G25'].value,  # G25
    "miri_formula_result": ws['G26'].value, # G26
}

# Prepare prompt for Ollama (Validation) - Focus Purely on Analysis Logic
# Removed any formatting instructions for the AI's output.
validation_prompt = f"""
Perform a reconciliation validation analysis for date {results_summary['date']}.

Input Data Summary:
- MC Summary Net Amount: {-net_amt_mc:.2f}
- MC Summary Net Count: {net_cnt_mc}
- MC DPR Amount: {results_summary['mc_dpr_amount']:.2f}
- MC DPR Count: {results_summary['mc_dpr_count']}
- MC Difference (Summary - DPR): {results_summary['mc_difference']:.2f}

- VISA Summary Net Amount: {-net_amt_visa:.2f}
- VISA Summary Net Count: {net_cnt_visa}
- VISA VSS Amount: {results_summary['visa_vss_amount']:.2f}
- VISA VSS Count: {results_summary['visa_vss_count']}
- VISA Difference (Summary - VSS): {results_summary['visa_difference']:.2f}

- MIR Summary Net Amount: {-net_amt_mir:.2f}
- MIR Summary Net Count: {net_cnt_mir}
- MIR ANR Amount: {results_summary['mir_anr_amount']:.2f}
- MIR ANR Count: {results_summary['mir_anr_count']}
- MIR Difference (Summary - ANR): {results_summary['mir_difference']:.2f}

NEW CLEARINGOPER Data:
- NMI Count (F24): {results_summary['nmi_count']}
- NVI Count (F25): {results_summary['nvi_count']}
- MIRI Count (F26): {results_summary['miri_count']}
- NMI Formula Result (G24): {results_summary['nmi_formula_result']}
- NVI Formula Result (G25): {results_summary['nvi_formula_result']}
- MIRI Formula Result (G26): {results_summary['miri_formula_result']}

Analysis Request:
1. Tolerance Check: Are all calculated differences (MC, VISA, MIR) within an acceptable tolerance of 0.01?
2. Discrepancy Identification: Highlight any discrepancies greater than 0.01.
3. Transaction Count: Compare record counts; flag mismatches > 1.
4. Gross & Net Validation: Verify gross amounts and recomputed net (gross - fees) match bank records.
5. Fee Consistency: Identify unexpected fee deviations.
6. Date & Currency: Confirm all entries are within settlement date and correct currency.
7. Data Integrity: Flag duplicates, missing IDs, or zero/negative anomalies.
8. CLEARINGOPER Analysis: Validate the new NMI, NVI, MIRI counts and their formulas.
9. Conclusion: Summarize as "All reconciled successfully" or "Discrepancies found: [list systems + issue types]".

"""

# Prepare prompt for Ollama (Simple Bar Chart Code) - Revised for Single Point
# Ask for a bar chart with distinct bars for each system.
chart_prompt = f"""
Generate Python code using matplotlib to create a bar chart for reconciliation differences.
Data Point: Date: {results_summary['date']}, MC Diff: {results_summary['mc_difference']:.2f}, VISA Diff: {results_summary['visa_difference']:.2f}, MIR Diff: {results_summary['mir_difference']:.2f}.
Also include the CLEARINGOPER data: NMI: {results_summary['nmi_formula_result']}, NVI: {results_summary['nvi_formula_result']}, MIRI: {results_summary['miri_formula_result']}.
Instructions:
1. Use the following exact structure for the code:
```python
import matplotlib.pyplot as plt

# Define the data as numeric values (not strings)
payment_systems = ['MC', 'VISA', 'MIR', 'NMI', 'NVI', 'MIRI']
# IMPORTANT: The differences must be numeric values (floats), NOT strings.
# Example: [1.5, 2.0, -0.5, 1.0, -0.2, 0.8] is correct.
differences = [{results_summary['mc_difference']:.2f}, {results_summary['visa_difference']:.2f}, {results_summary['mir_difference']:.2f}, {results_summary['nmi_formula_result']}, {results_summary['nvi_formula_result']}, {results_summary['miri_formula_result']}]

# Create the bar chart
plt.figure(figsize=(10, 6)) # Set figure size
plt.bar(payment_systems, differences, color=['red', 'blue', 'green', 'orange', 'purple', 'brown'], lw=2, alpha=0.7)

# Add labels and title
plt.xlabel('Payment System')
plt.ylabel('Difference')
plt.title('Reconciliation Health Dashboard - Including CLEARINGOPER Data')

# Add grid for better readability
plt.grid(True, linestyle='--', alpha=0.7)

# Adjust layout to prevent clipping
plt.tight_layout()

# Save the chart
plt.savefig('Result - recon_health_with_clearingoper_{results_summary['date']}.png')
"""

# === STEP 8: Query Ollama for Validation with Detailed Logging ===
print("🔍 Sending validation query to Ollama...")
print(f"   [DEBUG] URL: {OLLAMA_URL}")
print(f"   [DEBUG] Model: {OLLAMA_MODEL}")
# print(f"   [DEBUG] Prompt Length: {len(validation_prompt)} chars") # Optional: Log prompt size

try:
    print("   [REQUEST SENT] Validation query dispatched to Ollama.")
    response_validation = requests.post(
        OLLAMA_URL,
        json={
            "model": OLLAMA_MODEL,
            "prompt": validation_prompt,
            "stream": False
        },
        timeout=300 # Longer, simpler timeout
    )
    print(f"   [RESPONSE RECEIVED] Status Code: {response_validation.status_code}")
    print(f"   [DEBUG] Response Headers: {dict(response_validation.headers)}")
    # print(f"   [DEBUG] Response Text (first 200 chars): {response_validation.text[:200]}...") # Optional: Log part of response

    if response_validation.status_code == 200:
        ai_validation_result = response_validation.json().get("response", "No response received from Ollama for validation.")
        print("✅ Validation analysis received from Ollama.")
    else:
        ai_validation_result = f"Ollama returned status code: {response_validation.status_code}. Response: {response_validation.text}"
        print(f"❌ Ollama validation request failed with status {response_validation.status_code}.")

except requests.exceptions.Timeout:
    ai_validation_result = "Error: Ollama validation query timed out after 300 seconds."
    print("❌ Ollama validation query timed out.")
except requests.exceptions.ConnectionError:
    ai_validation_result = "Error: Could not connect to Ollama server. Is it running on http://localhost:11434?"
    print("❌ Connection error: Cannot reach Ollama server.")
except Exception as e:
    ai_validation_result = f"Error querying Ollama for validation: {e}"
    print(f"❌ Unexpected error querying Ollama for validation: {e}")

# === STEP 9: Query Ollama for Simple Bar Chart Code with Detailed Logging ===
print("🎨 Sending simple bar chart code generation query to Ollama...")
ai_chart_code = "No code received - Ollama request failed or timed out." # Default value
chart_generation_successful = False
print(f"   [DEBUG] URL: {OLLAMA_URL}")
print(f"   [DEBUG] Model: {OLLAMA_MODEL}")
# print(f"   [DEBUG] Prompt Length: {len(chart_prompt)} chars") # Optional: Log prompt size

try:
    print("   [REQUEST SENT] Bar chart code generation query dispatched to Ollama.")
    response_chart = requests.post(
        OLLAMA_URL,
        json={
            "model": OLLAMA_MODEL,
            "prompt": chart_prompt,
            "stream": False
        },
        timeout=300 # Longer, simpler timeout
    )
    print(f"   [RESPONSE RECEIVED] Status Code: {response_chart.status_code}")
    print(f"   [DEBUG] Response Headers: {dict(response_chart.headers)}")
    # print(f"   [DEBUG] Response Text (first 200 chars): {response_chart.text[:200]}...") # Optional: Log part of response

    if response_chart.status_code == 200:
        raw_ai_chart_code = response_chart.json().get("response", "No code received from Ollama for chart.")
        print("✅ Raw simple bar chart code received from Ollama.")

        # --- NEW: Extract Python code from Markdown block ---
        # Regex to find content between ```python and ```
        code_block_match = re.search(r'```python\s*\n(.*?)\n```', raw_ai_chart_code, re.DOTALL)
        if code_block_match:
            ai_chart_code = code_block_match.group(1)
            print("   [CODE EXTRACTED] Python code successfully parsed from Markdown block.")
            chart_generation_successful = True
        else:
            # If no markdown block found, use the whole response (might be plain code or an error)
            ai_chart_code = raw_ai_chart_code
            print("   [WARNING] No Markdown code block found in AI response. Using raw response as code.")
            # Consider this still successful if a response was received, even if parsing failed
            chart_generation_successful = True

        # --- END NEW SECTION ---

    else:
        ai_chart_code = f"Ollama returned status code: {response_chart.status_code}. Response: {response_chart.text}"
        print(f"❌ Ollama chart code request failed with status {response_chart.status_code}.")

except requests.exceptions.Timeout:
    ai_chart_code = "Error: Ollama chart code query timed out after 300 seconds."
    print("❌ Ollama chart code query timed out.")
    chart_generation_successful = False # Explicitly mark failure
except requests.exceptions.ConnectionError:
    ai_chart_code = "Error: Could not connect to Ollama server. Is it running on http://localhost:11434?"
    print("❌ Connection error: Cannot reach Ollama server.")
    chart_generation_successful = False # Explicitly mark failure
except Exception as e:
    ai_chart_code = f"Error querying Ollama for chart code: {e}"
    print(f"❌ Unexpected error querying Ollama for chart code: {e}")
    chart_generation_successful = False # Explicitly mark failure

# === STEP 10: Generate Python-Formatted Data Summary Block ===
# Define the width for the label part (left-aligned) and the value part (right-aligned)
label_width = 25
value_width = 25

def format_row(label, value, is_decimal=True):
    """Helper function to format a single row."""
    if is_decimal:
        value_str = f"{value:> {value_width}.2f}" # Right-align decimal number
    else:
        value_str = f"{value:> {value_width}}" # Right-align integer
    return f"{label:<{label_width}}{value_str}"

# Build the formatted summary block string using Python
formatted_summary_block = f"""System: MC
{format_row("Summary Net Amount:", -net_amt_mc)}
{format_row("Summary Net Count:", net_cnt_mc, is_decimal=False)}
{format_row("DPR/ANR Amount:", results_summary['mc_dpr_amount'])}
{format_row("DPR/ANR Count:", results_summary['mc_dpr_count'], is_decimal=False)}
{format_row("Calculated Diff:", results_summary['mc_difference'])}
System: VISA
{format_row("Summary Net Amount:", -net_amt_visa)}
{format_row("Summary Net Count:", net_cnt_visa, is_decimal=False)}
{format_row("DPR/ANR Amount:", results_summary['visa_vss_amount'])}
{format_row("DPR/ANR Count:", results_summary['visa_vss_count'], is_decimal=False)}
{format_row("Calculated Diff:", results_summary['visa_difference'])}
System: MIR
{format_row("Summary Net Amount:", -net_amt_mir)}
{format_row("Summary Net Count:", net_cnt_mir, is_decimal=False)}
{format_row("DPR/ANR Amount:", results_summary['mir_anr_amount'])}
{format_row("DPR/ANR Count:", results_summary['mir_anr_count'], is_decimal=False)}
{format_row("Calculated Diff:", results_summary['mir_difference'])}
NEW CLEARINGOPER Data:
{format_row("NMI Count (F24):", results_summary['nmi_count'], is_decimal=False)}
{format_row("NVI Count (F25):", results_summary['nvi_count'], is_decimal=False)}
{format_row("MIRI Count (F26):", results_summary['miri_count'], is_decimal=False)}
{format_row("NMI Formula Result (G24):", results_summary['nmi_formula_result'])}
{format_row("NVI Formula Result (G25):", results_summary['nvi_formula_result'])}
{format_row("MIRI Formula Result (G26):", results_summary['miri_formula_result'])}
"""

# === STEP 11: Save AI Validation Output (Combined Python Block + AI Analysis) ===
print("💾 Saving AI Validation output...")
with open(ai_validation_output_file, "w", encoding="utf-8") as f:
    f.write("=== AI Reconciliation Validation Report ===\n\n")
    f.write("--- FORMATTED DATA SUMMARY ---\n")
    f.write(formatted_summary_block) # Write the Python-generated, correctly formatted block first
    f.write("\n--- AI ANALYSIS ---\n")
    f.write(ai_validation_result) # Write the AI's analysis result (e.g., its conclusion)

print(f"📄 AI Validation Report saved: {ai_validation_output_file}")

# === STEP 11B: Save AI Chart Code Output (Only if successful) ===
if chart_generation_successful:
    print("💾 Saving AI Chart Code output...")
    with open(chart_code_output_file, "w", encoding="utf-8") as f:
        f.write("# Generated by Ollama AI (gemma3:4b)\n\n")
        f.write(ai_chart_code)
    print(f"📄 AI-Generated Chart Code saved: {chart_code_output_file}")
else:
    print("⚠️  Skipping saving AI Chart Code output due to error/time-out.")


# === STEP 12: Execute Generated Chart Code (Optional - Be Careful! - Only if successful) ===
if chart_generation_successful:
    execute_chart = input("\nDo you want to execute the AI-generated chart code now? (y/N): ").lower().strip()
    if execute_chart == 'y':
        print("🚀 Executing AI-generated chart code...")
        try:
            exec(ai_chart_code)
            print("✅ Chart generated successfully (if code was valid).")
        except Exception as e:
            print(f"❌ Error executing chart code: {e}")
            print("   The generated code might need manual review/editing.")
else:
    print("\n⚠️  Skipping chart code execution as generation failed or timed out.")

# --- MC ANR PROCESSING BEGINS HERE ---
# This part will now execute regardless of the chart code generation outcome.
print("\n--- MC ANR Processing Starting ---")

# Configuration for MC ANR
target_member_id_9176 = "00000009176"
associated_member_id_11682 = "00000011682"
anr_file_keyword = "NSPK-MC-ANR"

# Helper to find ANR file
def find_file_by_keyword(keyword):
    search_pattern = f"*{keyword}*"
    matching_files = glob.glob(search_pattern)
    matching_files = [f for f in matching_files if os.path.isfile(f)]
    if len(matching_files) == 0:
        raise FileNotFoundError(f"No file found containing keyword: {keyword}")
    elif len(matching_files) > 1:
        raise FileNotFoundError(f"Multiple files found containing keyword '{keyword}': {matching_files}. Please ensure uniqueness.")
    else:
        return matching_files[0]

try:
    anr_file = find_file_by_keyword(anr_file_keyword)
    print(f"🔍 Found ANR file: {anr_file}")
except FileNotFoundError as e:
    print(f"❌ Error: {e}")
    anr_file = None

if anr_file:
    # Attempt to read the file, handling potential encoding issues
    content = ""
    print(f"   - Attempting to read ANR file: {anr_file}")
    try:
        with open(anr_file, 'r', encoding='utf-8') as f:
            content = f.read()
        print("   - File read successfully with UTF-8 encoding.")
    except UnicodeDecodeError:
        print("   - UTF-8 decoding failed. Trying cp1252 encoding...")
        try:
            with open(anr_file, 'r', encoding='cp1252') as f:
                content = f.read()
            print("   - File read successfully with cp1252 encoding.")
        except UnicodeDecodeError:
            print(f"   - Failed to read file with UTF-8 or cp1252 encoding. Check file: {anr_file}")
            # Exit or handle the error as needed
            content = "" # Set to empty string to prevent further processing if read fails

    if content:
        # Split content by sections
        sections = re.split(r"CLEARING CYCLE 001 - NOTIFICATION", content)[1:]

        # Initialize variables
        subtotal_data_9176 = {}
        orig_data_11682 = {}
        count_11682 = 0

        # --- MAIN LOOP OVER SECTIONS ---
        for section in sections:
            # Identify Member ID for the section
            member_id_match = re.search(r"MEMBER ID:\s*(\d+)", section)
            if not member_id_match:
                continue # Skip sections without a clear member ID
            member_id = member_id_match.group(1)

            # --- Process Section for Target Member (9176) SUBTOTAL ---
            if member_id == target_member_id_9176:
                print(f"   - Found SUBTOTAL section for Target Member {target_member_id_9176}")
                # Find SUBTOTAL line within the section for the target member
                subtotal_match = re.search(r"SUBTOTAL\s+(\d+)\s+(\d+(?:,\d{3})*\.\d{2})\s+(DR|CR)\s+\d+-RUB\s+(\d+(?:,\d{3})*\.\d{2})\s+(DR|CR)", section)
                if subtotal_match:
                    count_str, recon_amount_str, recon_sign, fee_amount_str, fee_sign = subtotal_match.groups()
                    print(f"   - SUBTOTAL Match: Count='{count_str}', Recon='{recon_amount_str} {recon_sign}', Fee='{fee_amount_str} {fee_sign}'")
                    subtotal_data_9176 = {
                        "count": int(count_str.replace(',', '')),
                        "recon_amount": float(recon_amount_str.replace(',', '')),
                        "recon_sign": recon_sign,
                        "fee_amount": float(fee_amount_str.replace(',', '')),
                        "fee_sign": fee_sign
                    }

            # --- Process Section for Associated Member (11682) ORIG and SUBTOTAL ---
            if member_id == associated_member_id_11682:
                print(f"   - Found ORIG/SUBTOTAL section for Associated Member {associated_member_id_11682}")
                # Find the specific ORIG line that contributes opposing amounts for 9176
                # It should have the specific BUSINESS_SERVICE_ID and FILE ID related to 9176 within the 11682 section
                # Pattern: 643001 [FILE_ID] ORIG [AMOUNT1] CR [AMOUNT2] [SIGN2] (SIGN2 can be CR or DR)
                # Actual line examples:
                # 643001   021/260118/00000009176/01101 ORIG                17,159.33 CR                        50.33 CR  (Fee is CR)
                # 643001   021/260119/00000009176/01101 ORIG                64,252.00 CR                         0.00 DR  (Fee is DR)
                # 643001   021/260120/00000009176/01101 ORIG                52,276.02 CR                         0.00 DR  (Fee is DR) <- NEW FORMAT
                # Updated pattern to capture the fee amount and its sign, with corrected date regex
                orig_match = re.search(r"643001\s+021/2601\d{2}/00000009176/01101\s+ORIG\s+(\d+(?:,\d{3})*\.\d{2})\s+CR\s+(\d+(?:,\d{3})*\.\d{2})\s+(CR|DR)", section)
                if orig_match:
                    orig_recon_str, orig_fee_amount_str, orig_fee_sign = orig_match.groups()
                    orig_fee_amount_val = float(orig_fee_amount_str.replace(',', ''))
                    # Store the amount and its original sign
                    orig_data_11682 = {
                        "recon_amount": float(orig_recon_str.replace(',', '')),
                        "fee_amount": orig_fee_amount_val,
                        "fee_sign": orig_fee_sign
                    }
                else:
                    print(f"   - ORIG Match for 9176 data in 11682 section: NOT FOUND")

                # Find the SUBTOTAL line for member 11682 to get its count
                subtotal_11682_match = re.search(r"SUBTOTAL\s+(\d+)\s+", section)
                if subtotal_11682_match:
                    count_11682 = int(subtotal_11682_match.group(1).replace(',', ''))
                    print(f"   - SUBTOTAL Count for Member 11682: {count_11682}")

        # --- Perform Calculation (Following Explicit Instructions) ---
        # This calculation happens *after* the loop over sections
        if subtotal_data_9176 and orig_data_11682:
            # Apply the specific formula logic (Explicit Instructions)
            # 1. Take Primary DR amount (magnitude) from SUBTOTAL
            primary_dr_mag = subtotal_data_9176["recon_amount"]
            # 2. Take Opposing CR amount (magnitude) from ORIG
            opposing_cr_mag = orig_data_11682["recon_amount"]
            # 3. Calculate Recon Net Magnitude: Positive_Base_DR_Mag - Opposing_CR_Mag
            recon_net_magnitude = primary_dr_mag - opposing_cr_mag # e.g., 23,773,698.16 - 52,276.02

            # 4. Take Fee amounts (magnitudes and signs) from ORIG and SUBTOTAL
            subtotal_fee_mag = subtotal_data_9176["fee_amount"]
            subtotal_fee_sign = subtotal_data_9176["fee_sign"]
            orig_fee_mag = orig_data_11682["fee_amount"]
            orig_fee_sign = orig_data_11682["fee_sign"]

            # 5. Apply "reverse logic" for fees: if a fee's original sign is CR, add its magnitude. If DR, subtract its magnitude.
            subtotal_fee_contribution = subtotal_fee_mag if subtotal_fee_sign == "CR" else -subtotal_fee_mag
            orig_fee_contribution = orig_fee_mag if orig_fee_sign == "CR" else -orig_fee_mag
            fee_net_contribution = subtotal_fee_contribution + orig_fee_contribution # e.g., +376,139.93 + (-0.00) for file 20260120

            # 6. Calculate Final Amount: Recon Net Mag + Sum of Fee Contributions (using reverse logic)
            target_amount = recon_net_magnitude + fee_net_contribution # e.g., (23,773,698.16 - 52,276.02) + (376,139.93 - 0.00)

            # 7. Calculate Final Count: Sum of SUBTOTAL counts from both relevant members
            target_count = subtotal_data_9176["count"] + count_11682 # e.g., 19279 + 32

            print(f"   - Recon Net Calc: {primary_dr_mag} - {opposing_cr_mag} = {recon_net_magnitude}")
            print(f"   - Fee Net Calc: ({subtotal_fee_sign} {subtotal_fee_mag}) + ({orig_fee_sign} {orig_fee_mag}) -> Rev. Logic: (+/-){subtotal_fee_contribution} + (+/-){orig_fee_contribution} = {fee_net_contribution}")
            print(f"   - Final Amount Calc: {recon_net_magnitude} + {fee_net_contribution} = {target_amount}")
            print(f"   - Final Count Calc: {subtotal_data_9176['count']} + {count_11682} = {target_count}")

            # Update the Excel worksheet object (wb/ws) with the new E22 and F22 values
            # Apply number formatting to E22 and F22
            ws['E22'].number_format = '#,##0.00'
            ws['F22'].number_format = '#,##0'

            # Place the calculated values into the worksheet
            ws['E22'] = target_amount  # E22 (Amount)
            ws['F22'] = target_count   # F22 (Count)

            print(f"   - E22 (MC ANR Amount) updated to: {target_amount:.2f}")
            print(f"   - F22 (MC ANR Count) updated to: {target_count}")

        else:
            print("❌ Required data (SUBTOTAL for 9176 or ORIG from 11682) not found in the file sections.")
            print("   - MC ANR row (E22, F22) will not be updated.")
    else:
        print(f"❌ Could not read content of ANR file: {anr_file}. Skipping MC ANR processing.")
else:
    print("❌ Could not find ANR file. MC ANR processing skipped.")
    print("   - MC ANR row (E22, F22) will not be updated.")

# --- G22 FORMULA ADDITION ---
print("\n--- Adding G22 Formula ---")
# Add the formula to G22 (Row 22, Col 7 -> G)
# This formula checks reconciliation between MC DPR (E19) and MC ANR (E22)
ws['G22'] = f'=E19+E22' # Add the formula here
ws['G22'].number_format = '#,##0.00' # Apply number format to G22
print(f"   - Formula '=E19+E22' added to cell G22.")

# --- SAVE FINAL WORKBOOK ---
wb.save(output_file)
wb.close()
os.remove(temp_file) # Clean up the temporary file

print(f"\n✅ Reconciliation Script v37.1 completed!")
print(f"📁 Final output saved: {output_file}")
print("--- AI Validation Complete ---")
if chart_generation_successful:
    print("--- AI Chart Code Generation Complete ---")
else:
    print("--- AI Chart Code Generation Failed/Skipped ---")
print("--- MC ANR Integration Complete ---")
print("--- G22 Formula Added ---")
print("--- CLEARINGOPER Integration Complete ---")
print("🎉 Full Process Complete!")